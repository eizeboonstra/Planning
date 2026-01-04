from gurobipy import *
from openpyxl import *
from time import *
import math
##
class Airport:
    def __init__(self, id,latitude, longitude, runway, slots_available):
        self.id  = id
        self.lat = latitude
        self.long = longitude
        self.Runway = runway
        self.slots = slots_available
        self.InLinks  = [ ]         
        self.OutLinks = [ ]        
    
    def addInLink(self,arc):      # Add new 'In Link' to the node
        self.InLinks.append(arc)       
    
    def addOutLink(self,arc):     # Add new 'Out Link' to the node
        self.OutLinks.append(arc)

class Arc:
    def __init__(self,origin, destination,demand,distance):
        self.From   = origin      #i
        self.To     = destination #j
        self.Dist   = distance  #km
        self.Dem  = demand

class Aircraft:
    def __init__ (self,id,speed,seats,TAT,Range,Runway,leasing_costs,fixed_costs,timebased_costs,fuel_parameter_costs):
        self.id = id
        self.s = seats
        self.C_L = leasing_costs #weekly
        self.C_X = fixed_costs #per flight
        self.c_T = timebased_costs #per hour
        self.c_F = fuel_parameter_costs #per distance
        self.V = speed
        self.R = Range
        self.TAT = TAT
        self.Runway = Runway
##

def read_matrix(file_name, sheet_name):
    wb = load_workbook(file_name, data_only=True)
    ws = wb[sheet_name]

    # Row 1: Airport, PDL, LIS, OPO, ...
    header = [c for c in next(ws.iter_rows(min_row=1, values_only=True))]
    airports = header[1:]  # PDL, LIS, OPO, ...
    
    data = []
    id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        origin = row[0]
        values = row[1:]

        for dest, val in zip(airports, values):
            if origin == dest:
                continue
            if val is None or val == 0:
                continue
            data.append((id,origin,dest,val))
            id +=1

    return data


def read_file(file_name,sheet_name):

    wb = load_workbook(file_name, data_only=True)
    ws = wb[sheet_name]
    List = []
    for cols in ws.iter_cols(min_col=2, values_only = True):
        values = cols[0:]
        List.append((values))
    return List



def read_arcs_from_file(file, sheet_dist, sheet_dem):
    """
    Crea List_Arcs = [(id, origin, dest, distance, demand), ...]
    usando:
      - tabella distanze
      - tabella domande
    Crea List_Arcs = [(id, origin, dest, distance, demand), ...]
    usando:
      - tabella distanze
      - tabella domande
    """
    dist_dict = read_matrix(file, sheet_dist)
    dem_dict  = read_matrix(file,  sheet_dem)

    List_Arcs = []
    arc_id = 1

    for (origin, dest), d_km in dist_dict.items():
        dem = dem_dict.get((origin, dest), 0)  # 0 se manca
        List_Arcs.append((arc_id, origin, dest, d_km, dem))
        arc_id += 1

    return List_Arcs

def calculate_distance (a,airport):
    R_e = 6371 #[km]
    k = 0
    for ap in airport:
            if a[1] == ap.id:
                k+=1
                lam_from = math.radians(ap.long)
                phi_from = math.radians(ap.lat)
            elif a[2] == ap.id:
                k+=1
                lam_to = math.radians(ap.long)
                phi_to = math.radians(ap.lat)
            if k == 2:
                break
        
    delta = 2*math.asin(math.sqrt(math.sin((phi_from-phi_to)/2)**2+math.cos(phi_from)*math.cos(phi_to)*(math.sin((lam_from-lam_to)/2))**2))
    dist_a = delta*R_e
    return (dist_a)

def construct_graph():
    arc = []
    airport = []
    aircraft = []
    airport_map={}
    # (ID, origin, destination, distance_km, demand_pax)
    List_Arcs = read_matrix ("airport_1b.xlsx","Demands")
    List_Airports = read_file("airport_1b.xlsx","Airports")

    for (id,latitude, longitude, runway, slots_available) in List_Airports:
        ap = Airport(id,latitude, longitude, runway, slots_available)
        airport.append(ap)
        airport_map[id] = ap
    
    #generare la distanza
    #generare la distanza
    for i,a in enumerate(List_Arcs):
        a = a + (calculate_distance(a, airport),)
        List_Arcs[i] = a
    for (id,origin,destination,demand,distance) in List_Arcs:
        a = Arc(origin,destination,int(demand),distance)
        arc.append(a)
        airport_map[origin].addOutLink(a)
        airport_map[destination].addInLink(a)


    List_Aircraft = read_file("AircraftData.xlsx","Aircraft Data")
    for (id,speed,seats,TAT,Range,Runway,leasing_costs,fixed_costs,timebased_costs,fuel_parameter_costs) in List_Aircraft:
        aircraft.append(Aircraft(id,speed,seats,TAT,Range,Runway,leasing_costs,fixed_costs,timebased_costs,fuel_parameter_costs))
    return (arc,airport,aircraft)

def cost_definition(aircraft,arc):
    f = 1.42
    C_X = aircraft.C_X
    C_T = aircraft.c_T*arc.Dist/aircraft.V
    C_F =((aircraft.c_F*f)/1.5)*arc.Dist
    op_costs = C_X+C_T+C_F # per frequency # we apply a 30% reduction on the flight to the hub
    op_costs = 0.7*op_costs
    return op_costs

def yield_definition (arc):
    Yield = 5.9*arc.Dist**(-0.76)+0.043
    return Yield


#check if optimization was successful
def check_model_status(model):
    status = model.status
    if status != GRB.Status.OPTIMAL:
        if status == GRB.Status.UNBOUNDED:
            print('The model cannot be solved because it is unbounded')
        elif status == GRB.Status.INFEASIBLE:
            print('The model is infeasible; computing IIS')
            model.computeIIS()
            print('The following constraint(s) cannot be satisfied:')
            for c in model.getConstrs():
                if c.IISConstr:
                    print(c.constrName)
        elif status != GRB.Status.INF_OR_UNBD:
            print('Optimization was stopped with status',status)
        exit(0)


#print the result of the optimization into a readable text format




##########MODELLO GURIBI
##########MODELLO GURIBI
def Model_2 (arc,airport,aircraft):
    "Solve the MCF Problem with a Arc-based formulation using linear programming."
    LF = 0.75
    hub_airport = "EHAM"
    BT = 10
    #CASK = C_x + C_T + C_F with C_T = c_T*dij/Vk
    Budget = 10000000000000000000
    Budget = 10000000000000000000

    g = {}
    for ap in airport:
        if ap.id == hub_airport:
            g[ap.id] = 0
        else:
            g[ap.id] = 1
    # LP model (this is an object)
    model = Model("MCF") 
    x =  {}
    z = {}
    w = {}
    n_ac = {}
    for a in arc:
        w[a] = model.addVar(vtype =GRB.CONTINUOUS, name = ''.join(['hub_pax(', str(a.From), ',', str(a.To), ')']))
        if a.From == hub_airport or a.To == hub_airport:
            x[a] = model.addVar(vtype =GRB.CONTINUOUS, name = ''.join(['direct_pax(', str(a.From), ',', str(a.To), ')']))
            for ac in aircraft:
                z[a,ac] = model.addVar(vtype =GRB.INTEGER, name = ''.join(['freq(', str(a.From), ',', str(a.To), ',', str(ac.id), ')']))
    for ac in aircraft:
        n_ac[ac] = model.addVar(vtype =GRB.INTEGER, name = ''.join(['number of aircarft(', str(ac.id),')']))
    model.update()
    model.setObjective(
        quicksum(
            yield_definition(a)*a.Dist*(x.get(a,0)+w[a])-quicksum(cost_definition(ac,a)*z.get((a,ac),0) for ac in aircraft)
            for a in arc) - quicksum(n_ac[ac]*ac.C_L for ac in aircraft),
        GRB.MAXIMIZE
        )
    model.update()

    # cdemand constraints
    Demand_cr = {}
    for a in arc:
        Demand_cr[a.From,a.To] = model.addConstr((x.get(a,0)+w[a])<= a.Dem,
                                                    name = f"demand_cnstr{a.From}_{a.To}")

    # capacity constraint
    Capacity = {}
    for a in arc:
        term_out = quicksum(w[aw] for aw in arc if aw.From == a.From)*(1-g[a.To])
        term_in = quicksum(w[aw] for aw in arc if aw.To == a.To)*(1-g[a.From])
        Capacity[a.From,a.To] = model.addConstr(x.get(a,0)+term_out+term_in<=quicksum(z.get((a,ac),0)*LF*ac.s for ac in aircraft),name = f"capacity_cnstr{a.From}_{a.To}")


    #hub constraint
    Hub = {}
    for a in arc:
        Hub[a.From, a.To] = model.addConstr(w[a]<=a.Dem*g[a.From]*g[a.To], name =f"hub{a.From}_{a.To}")

    #balance constraint
    Balance = {}
    for ap in airport:
        for ac in aircraft:
            out_freq = quicksum(z.get((a,ac),0) for a in arc if a.From == ap.id)
            in_freq = quicksum(z.get((a,ac),0) for a in arc if a.To == ap.id)
            Balance[ap.id, ac.id] = model.addConstr(out_freq == in_freq, name =f"balance{ap.id}_{ac.id}")
    
    #time constraints
    Time = {}
    for ac in aircraft:
        Time = model.addConstr(quicksum(((1.5 if a.To == hub_airport else 1.0)*ac.TAT/60+a.Dist/ac.V)*z.get((a,ac),0) for a in arc)<=BT*7*n_ac[ac], name =f"time{ac.id}")
        Time = model.addConstr(quicksum(((1.5 if a.To == hub_airport else 1.0)*ac.TAT/60+a.Dist/ac.V)*z.get((a,ac),0) for a in arc)<=BT*7*n_ac[ac], name =f"time{ac.id}")
        
    #range constraint
    Range = {}
    for ac in aircraft:    
        for a in arc:
            if a.Dist <= ac.R:
                max_freq = 1000000
            else:
                max_freq = 0
            Range[a.From,a.To,ac.id] = model.addConstr(z.get((a,ac),0)<=max_freq,name =f"range {a.From}-{a.To} with {ac.id}")
    
    #cost constraint
    Cost= model.addConstr(quicksum(n_ac[ac]*ac.C_L for ac in aircraft)<=Budget, name =f"budget")

    #slot constraint
    Slot = {}
    for ap in airport:
        if ap.slots == '-':
            continue
        else:
            Slot[ap.id]= model.addConstr(quicksum(z.get((a,ac),0) for a in arc for ac in aircraft if a.To == ap.id)<=ap.slots, name =f"slot_constarint{ap.id}")

    #runway constraint
    Runway = {}
    for a in arc:
        for ac in aircraft:
            if z.get((a, ac), None) == None:
                continue
            else:
                k = 0
                for ap in airport:
                    if ap.id == a.To:
                        k += 1
                        if ac.Runway>ap.Runway:
                            Runway[a.From,a.To,ac.id]= model.addConstr(z[a,ac]==0, name =f"runway {a.From}-{a.To} with {ac.id} to {ap.id}")
                    elif ap.id == a.From:
                        k += 1
                        if ac.Runway>ap.Runway:
                            Runway[a.From,a.To,ac.id]= model.addConstr(z[a,ac]==0, name =f"runway {a.From}-{a.To} with {ac.id} from {ap.id}")
                    if k == 2:
                        break


    
    #save info to log file
    model.setParam("LogFile", 'log_file')
    #update gurobi with the constraints
    model.update()
    # Useful for model debugging
    model.write("Model_1.lp")
    model.write("Model_1.lp")
    model.optimize()
    
    check_model_status(model)
    print_model_result(x, w, z, n_ac, model, arc, aircraft, BT, hub_airport)



def print_model_result(x, w, z, n_ac, model, arc, aircraft, BT, hub_airport):
    """
    Stampa:
    - per ogni arco usato: pax totali e voli per tipo di aeromobile
    - valore dell'OF
    - numero di aeromobili per tipo
    - utilizzazione della flotta per tipo
    """

    if model.SolCount == 0:
        print("Nessuna soluzione trovata.")
        return

    print("Soluzione ottima (pax; voli per tipo di aeromobile):\n")

    # intestazione per ricordare l'ordine dei tipi di aereo
    ac_ids = [ac.id for ac in aircraft]
    print("Ordine tipi AC:", ", ".join(ac_ids), "\n")

    # --- Flussi di leg + voli per tipo ---
    for a in arc:
        # x[a] esiste solo se l’arco tocca l’hub
        x_var = x.get(a, None)
        x_val = x_var.X if x_var is not None else 0.0

        w_val = w[a].X
        pax_tot = x_val + w_val

        # voli per tipo
        flights = []
        for ac in aircraft:
            z_var = z.get((a, ac), None)
            if z_var is not None:
                flights.append(int(round(z_var.X)))
            else:
                flights.append(0)

        # se non ci sono pax e non ci sono voli, salto
        if pax_tot <= 1e-6 and all(f == 0 for f in flights):
            continue

        flights_str = ";".join(str(f) for f in flights)
        print(f"{a.From} -> {a.To}: {int(round(pax_tot))} pax  ({flights_str})")

    # --- Objective function ---
    print(f"\nOF = {model.ObjVal:.2f} €/week\n")

    # --- Fleet size ---
    print("Numero di aeromobili per tipo:")
    for ac in aircraft:
        n_val = n_ac[ac].X
        print(f"  {ac.id}: {n_val:.2f}")

    # --- Utilizzazione flotta (stime coerenti col vincolo di tempo) ---
    print("\nAC Utilisation (stime):")
    for ac in aircraft:
        used_hours = 0.0
        for a in arc:
            z_var = z.get((a, ac), None)
            if z_var is None or z_var.X <= 1e-6:
                continue

            # stesso tempo usato nel vincolo di tempo
            turn_time = (1.5 if a.To == hub_airport else 1.0) * ac.TAT / 60.0
            flight_time = a.Dist / ac.V
            used_hours += (turn_time + flight_time) * z_var.X

        n_val = n_ac[ac].X
        available_hours = BT * 7 * n_val  # 10 h/giorno * 7 * n_ac

        if available_hours > 1e-6:
            util = 100.0 * used_hours / available_hours
        else:
            util = 0.0

        print(f"  {ac.id}: {util:.2f}%")


###### RUNNARE IL CODICE

arc, airport, aircraft = construct_graph()

# RUN MCF PROBLEM
start_time = time()
Model_2(arc, airport,aircraft)    
print ('Run Time =', time() - start_time)

